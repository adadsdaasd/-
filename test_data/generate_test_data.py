# -*- coding: utf-8 -*-
"""
Generate test data files (Excel format)
Run: python generate_test_data.py
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os

# Ensure running in test_data directory
script_dir = os.path.dirname(os.path.abspath(__file__))
os.chdir(script_dir)

print("Generating test data files...")

# ==================== 1. Team members table (5 people) ====================
# All required fields: name, phone, education, work experience, skills
team_data = {
    "姓名": ["李明", "王芳", "张伟", "陈静", "刘洋"],
    "电话": ["13812345678", "13611112222", "13987654321", "13722223333", "13833334444"],
    "邮箱": ["liming@example.com", "wangfang@tencent.com", "zhangwei@cas.cn", "chenjing@sensetime.com", "liuyang@bytedance.com"],
    "学历": ["硕士", "本科", "博士", "硕士", "硕士"],
    "学校": ["清华大学", "北京大学", "中国科学院", "清华大学", "浙江大学"],
    "专业": ["计算机科学与技术", "软件工程", "人工智能", "计算机视觉", "软件工程"],
    "工作经历": [
        "阿里巴巴高级算法工程师(2019-至今)",
        "腾讯前端开发工程师(2021-至今)",
        "华为AI研究院副教授(2018-至今)",
        "商汤科技高级研究员(2020-至今)",
        "字节跳动后端架构师(2020-至今)"
    ],
    "工作年限": ["5年", "3年", "8年", "4年", "6年"],
    "技能特长": [
        "Python, 机器学习, 推荐系统, TensorFlow, Spark",
        "React, Vue, TypeScript, 微信小程序, CSS",
        "深度学习, NLP, PyTorch, 论文写作, 科研项目管理",
        "计算机视觉, PyTorch, 模型部署, TensorRT, ONNX",
        "Go, Java, 分布式系统, Kubernetes, Docker"
    ],
    "个人优势": [
        "扎实的算法基础和丰富的工业界经验，推荐系统领域专家",
        "精通现代前端框架，良好的用户体验意识，快速学习能力强",
        "NLP领域深厚积累，高水平论文发表能力，科研项目管理经验丰富",
        "CV领域实战经验丰富，多项发明专利，模型落地能力强",
        "分布式系统架构经验丰富，高并发处理能力强，技术视野开阔"
    ],
    "未来规划": [
        "成为AI算法领域技术专家，带领更大团队",
        "成为前端技术专家并探索全栈发展",
        "在学术领域继续深耕并培养优秀学生",
        "推动CV技术在工业界的大规模落地应用",
        "成为技术架构专家并参与开源社区建设"
    ],
    "当前绩效": [85.0, 72.5, 95.0, 88.0, 90.5],
    "主要贡献": [
        "主导推荐系统V2.0上线|20; 发表ACM论文1篇|10",
        "重构前端组件库|15; 完成小程序性能优化|8",
        "获批国家自然科学基金|30; 发表Nature子刊论文|25; 指导3名博士生|10",
        "TensorRT模型加速方案落地|18; 申请发明专利3项|12",
        "设计微服务架构迁移方案|25; 开源Go框架Star 2k+|15; 优化集群成本30%|10"
    ]
}

df_team = pd.DataFrame(team_data)
df_team.to_excel("team_5_members.xlsx", index=False, engine='openpyxl')
print("[OK] Generated: team_5_members.xlsx")


# ==================== 2. Blank job application form (single person) ====================
wb = Workbook()
ws = wb.active
ws.title = "求职申请表"

# Styles
label_font = Font(bold=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
gray_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

# Title
ws.merge_cells('A1:D1')
ws['A1'] = "求职申请表"
ws['A1'].font = Font(bold=True, size=18)
ws['A1'].alignment = Alignment(horizontal='center')

# Form fields (all required fields included)
fields = [
    ("姓名", ""),
    ("电话", ""),
    ("邮箱", ""),
    ("学历", ""),
    ("学校", ""),
    ("专业", ""),
    ("工作经历", ""),
    ("工作年限", ""),
    ("技能特长", ""),
    ("个人优势", ""),
    ("未来规划", ""),
]

row = 3
for label, value in fields:
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=1).font = label_font
    ws.cell(row=row, column=1).fill = gray_fill
    ws.cell(row=row, column=1).border = thin_border
    
    ws.merge_cells(f'B{row}:D{row}')
    ws.cell(row=row, column=2, value=value)
    ws.cell(row=row, column=2).border = thin_border
    
    if label in ["工作经历", "技能特长", "个人优势", "未来规划"]:
        ws.row_dimensions[row].height = 60
    
    row += 1

ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20

wb.save("blank_job_application.xlsx")
print("[OK] Generated: blank_job_application.xlsx")


# ==================== 3. Blank team summary table (multi-person) ====================
wb2 = Workbook()
ws2 = wb2.active
ws2.title = "团队信息汇总"

ws2.merge_cells('A1:N1')
ws2['A1'] = "团队成员信息汇总表"
ws2['A1'].font = Font(bold=True, size=16)
ws2['A1'].alignment = Alignment(horizontal='center')

# Headers - all required fields + performance/contribution
headers = ["序号", "姓名", "电话", "邮箱", "学历", "学校", "工作经历", "工作年限", "技能特长", "个人优势", "未来规划", "当前绩效", "主要贡献", "备注"]
for col, header in enumerate(headers, 1):
    cell = ws2.cell(row=3, column=col, value=header)
    cell.font = label_font
    cell.fill = gray_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

for row in range(4, 9):
    ws2.cell(row=row, column=1, value=row-3)
    for col in range(1, len(headers) + 1):
        ws2.cell(row=row, column=col).border = thin_border

col_widths = [6, 10, 15, 22, 8, 12, 25, 10, 25, 25, 20, 10, 30, 12]
for i, width in enumerate(col_widths, 1):
    ws2.column_dimensions[chr(64+i)].width = width

wb2.save("blank_team_summary.xlsx")
print("[OK] Generated: blank_team_summary.xlsx")


# ==================== 4. Blank research project proposal ====================
wb3 = Workbook()
ws3 = wb3.active
ws3.title = "项目申报书"

ws3.merge_cells('A1:D1')
ws3['A1'] = "科研项目申报书"
ws3['A1'].font = Font(bold=True, size=18)
ws3['A1'].alignment = Alignment(horizontal='center')

basic_fields = [
    ("项目名称", ""),
    ("项目负责人", ""),
    ("联系电话", ""),
    ("电子邮箱", ""),
    ("所在单位", ""),
    ("最高学历", ""),
    ("研究方向", ""),
    ("工作年限", ""),
    ("技能特长", ""),
    ("项目类型", ""),
    ("申请经费(万元)", ""),
    ("项目周期", ""),
]

ws3.merge_cells('A3:D3')
ws3['A3'] = "一、基本信息"
ws3['A3'].font = Font(bold=True, size=12)

row = 4
for label, value in basic_fields:
    ws3.cell(row=row, column=1, value=label)
    ws3.cell(row=row, column=1).font = label_font
    ws3.cell(row=row, column=1).fill = gray_fill
    ws3.cell(row=row, column=1).border = thin_border
    
    ws3.merge_cells(f'B{row}:D{row}')
    ws3.cell(row=row, column=2).border = thin_border
    row += 1

content_fields = [
    ("项目摘要", 80),
    ("研究背景与意义", 100),
    ("研究内容与目标", 100),
    ("研究方法与技术路线", 100),
    ("预期成果", 60),
    ("经费预算说明", 60),
]

row += 1
ws3.merge_cells(f'A{row}:D{row}')
ws3[f'A{row}'] = "二、项目内容"
ws3[f'A{row}'].font = Font(bold=True, size=12)
row += 1

for label, height in content_fields:
    ws3.cell(row=row, column=1, value=label)
    ws3.cell(row=row, column=1).font = label_font
    ws3.cell(row=row, column=1).fill = gray_fill
    ws3.cell(row=row, column=1).border = thin_border
    
    ws3.merge_cells(f'B{row}:D{row}')
    ws3.cell(row=row, column=2).border = thin_border
    ws3.row_dimensions[row].height = height
    row += 1

ws3.column_dimensions['A'].width = 22
ws3.column_dimensions['B'].width = 25
ws3.column_dimensions['C'].width = 25
ws3.column_dimensions['D'].width = 25

wb3.save("blank_research_proposal.xlsx")
print("[OK] Generated: blank_research_proposal.xlsx")


print("\nAll test data files generated successfully!")
print("\nGenerated files:")
print("  - team_5_members.xlsx (5 people with all required fields + performance + contributions)")
print("  - blank_job_application.xlsx (single person form)")
print("  - blank_team_summary.xlsx (multi-person summary with performance/contribution columns)")
print("  - blank_research_proposal.xlsx (research proposal)")
